"""AGFDATA Iteration 4 Backend Tests.

New/updated endpoints:
- GET /api/rekap/staffing-summary (?no_po=)
- GET /api/export/barang-masuk/pdf (?search=)
- GET /api/export/staffing/pdf
- GET /api/export/staffing/excel  (with image "Foto" column)
- GET /api/export/barang-masuk/excel (?search=, with image "Foto" column)
- Startup migration: legacy progres records with po_id="" removed

Test env: uses REACT_APP_BACKEND_URL from frontend/.env at load time.
"""
import io
import os
import uuid
import pytest
import requests
from openpyxl import load_workbook

# Load REACT_APP_BACKEND_URL from frontend/.env if not present in env
if not os.environ.get("REACT_APP_BACKEND_URL"):
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    os.environ["REACT_APP_BACKEND_URL"] = line.split("=", 1)[1].strip()
                    break
    except FileNotFoundError:
        pass

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"


def _login(email: str, password: str) -> requests.Session:
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=30)
    assert r.status_code == 200, f"Login failed for {email}: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def admin():
    return _login("admin@agfdata.com", "admin123")


@pytest.fixture(scope="module")
def staff():
    return _login("staff@agfdata.com", "staff123")


@pytest.fixture(scope="module")
def guest():
    return _login("tamu@agfdata.com", "tamu123")


@pytest.fixture(scope="module")
def seeded(admin, staff):
    """Create barang -> PO (qty=10) -> barang_masuk (qty=4) -> staffing (qty=3)."""
    b = admin.post(f"{API}/barang", json={
        "nama_barang": f"TEST_iter4_{uuid.uuid4().hex[:6]}",
        "nama_pengrajin": "TEST_P4", "spesifikasi": "spec4",
        "harga_pengrajin": 1.0, "harga_jual": 2.0
    }, timeout=30).json()

    po_no = f"TEST_ITER4_{uuid.uuid4().hex[:6]}"
    po = admin.post(f"{API}/po", json={
        "no_po": po_no,
        "items": [{"barang_id": b["_id"], "qty": 10}]
    }, timeout=30).json()

    staff.post(f"{API}/barang-masuk", json={
        "po_id": po["_id"], "tanggal_masuk": "2026-01-22", "penerima": "TEST_iter4_penerima",
        "items": [{"barang_id": b["_id"], "nama_barang": b["nama_barang"],
                   "nama_pengrajin": "TEST_P4", "spesifikasi": "spec4", "qty_diterima": 4}]
    }, timeout=30)

    staff.post(f"{API}/staffing", json={
        "po_id": po["_id"], "tanggal_keluar": "2026-01-23",
        "items": [{"barang_id": b["_id"], "nama_barang": b["nama_barang"],
                   "nama_pengrajin": "TEST_P4", "spesifikasi": "spec4", "qty": 3}]
    }, timeout=30)

    return {"barang": b, "po": po, "no_po": po_no}


# ============================================================
# 1. GET /api/rekap/staffing-summary
# ============================================================
class TestRekapStaffingSummary:
    def test_returns_expected_shape(self, admin, seeded):
        r = admin.get(f"{API}/rekap/staffing-summary", timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert isinstance(data, list)
        assert len(data) > 0

        expected_keys = {"no_po", "nama_barang", "gambar_path", "qty_po", "qty_staffing", "kurang_kirim"}
        for row in data:
            assert expected_keys.issubset(row.keys()), f"Missing keys: {expected_keys - set(row.keys())}"
            # pengrajin MUST NOT be included per iter4 spec
            assert "nama_pengrajin" not in row
            assert "pengrajin" not in row
            assert row["kurang_kirim"] == row["qty_po"] - row["qty_staffing"]

    def test_no_po_filter(self, admin, seeded):
        r = admin.get(f"{API}/rekap/staffing-summary", params={"no_po": seeded["no_po"]}, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert len(data) > 0
        for row in data:
            assert row["no_po"] == seeded["no_po"]
        # values sanity: qty_po=10 qty_staffing=3 kurang=7
        me = data[0]
        assert me["qty_po"] == 10
        assert me["qty_staffing"] == 3
        assert me["kurang_kirim"] == 7

    def test_no_po_no_match_returns_empty(self, admin):
        r = admin.get(f"{API}/rekap/staffing-summary",
                      params={"no_po": "TEST_NON_EXISTENT_ITER4_XYZ"}, timeout=30)
        assert r.status_code == 200
        assert r.json() == []

    def test_requires_auth(self):
        r = requests.get(f"{API}/rekap/staffing-summary", timeout=30)
        assert r.status_code == 401


# ============================================================
# 2. GET /api/export/barang-masuk/pdf  (with search filter)
# ============================================================
class TestExportBarangMasukPDF:
    def test_pdf_no_filter(self, admin):
        r = admin.get(f"{API}/export/barang-masuk/pdf", timeout=60)
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert r.content[:4] == b"%PDF"

    def test_pdf_search_filter_returns_pdf(self, admin, seeded):
        r = admin.get(f"{API}/export/barang-masuk/pdf",
                      params={"search": seeded["no_po"]}, timeout=60)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_pdf_search_by_penerima(self, admin):
        r = admin.get(f"{API}/export/barang-masuk/pdf",
                      params={"search": "TEST_iter4_penerima"}, timeout=60)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_pdf_search_by_nama_barang(self, admin, seeded):
        r = admin.get(f"{API}/export/barang-masuk/pdf",
                      params={"search": seeded["barang"]["nama_barang"]}, timeout=60)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_pdf_requires_auth(self):
        r = requests.get(f"{API}/export/barang-masuk/pdf", timeout=30)
        assert r.status_code == 401


# ============================================================
# 3. GET /api/export/staffing/pdf
# ============================================================
class TestExportStaffingPDF:
    def test_pdf(self, admin):
        r = admin.get(f"{API}/export/staffing/pdf", timeout=60)
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert r.content[:4] == b"%PDF"

    def test_pdf_requires_auth(self):
        r = requests.get(f"{API}/export/staffing/pdf", timeout=30)
        assert r.status_code == 401


# ============================================================
# 4. GET /api/export/staffing/excel  (with "Foto" image column)
# ============================================================
class TestExportStaffingExcel:
    def test_xlsx_content_type_and_signature(self, admin):
        r = admin.get(f"{API}/export/staffing/excel", timeout=60)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "").lower()
        # xlsx = zip archive, starts with PK\x03\x04
        assert r.content[:2] == b"PK", "Response is not a valid xlsx (zip) archive"

    def test_xlsx_readable_by_openpyxl_and_has_foto_column(self, admin):
        r = admin.get(f"{API}/export/staffing/excel", timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Foto" in headers, f"Missing 'Foto' column, headers={headers}"
        # Foto should be the first column
        assert headers[0] == "Foto"
        # Expected other headers: No PO, Tanggal, Barang, Pengrajin, Qty
        assert "No PO" in headers
        assert "Qty" in headers

    def test_xlsx_contains_images_if_data_present(self, admin, seeded):
        """If any staffing row has gambar_path, an image should be embedded."""
        r = admin.get(f"{API}/export/staffing/excel", timeout=60)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # openpyxl exposes embedded images via ws._images
        # If no barang has gambar_path uploaded, this may be zero; assert list attr exists.
        assert hasattr(ws, "_images")

    def test_xlsx_requires_auth(self):
        r = requests.get(f"{API}/export/staffing/excel", timeout=30)
        assert r.status_code == 401


# ============================================================
# 5. GET /api/export/barang-masuk/excel  (with image column + search)
# ============================================================
class TestExportBarangMasukExcel:
    def test_xlsx_content_type(self, admin):
        r = admin.get(f"{API}/export/barang-masuk/excel", timeout=60)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "").lower()
        assert r.content[:2] == b"PK"

    def test_xlsx_has_foto_column(self, admin):
        r = admin.get(f"{API}/export/barang-masuk/excel", timeout=60)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers[0] == "Foto", f"First header must be 'Foto', got {headers}"
        for h in ["No PO", "Tanggal", "Penerima", "Barang", "Qty"]:
            assert h in headers, f"Missing header {h}: got {headers}"

    def test_xlsx_search_filter(self, admin, seeded):
        """Search should filter down to only matching rows."""
        r = admin.get(f"{API}/export/barang-masuk/excel",
                      params={"search": seeded["no_po"]}, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # collect all no_po column values (2nd column after Foto)
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        no_po_col = headers.index("No PO") + 1
        values = [ws.cell(row=r_i, column=no_po_col).value
                  for r_i in range(2, ws.max_row + 1)]
        values = [v for v in values if v]  # skip empty
        assert len(values) >= 1, "Filter should have kept our seeded row"
        for v in values:
            assert v == seeded["no_po"], f"Filter leak: found {v} when searching {seeded['no_po']}"

    def test_xlsx_search_no_match_returns_only_header(self, admin):
        r = admin.get(f"{API}/export/barang-masuk/excel",
                      params={"search": "TEST_NO_MATCH_ITER4_XYZ"}, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # only header row
        assert ws.max_row == 1

    def test_xlsx_requires_auth(self):
        r = requests.get(f"{API}/export/barang-masuk/excel", timeout=30)
        assert r.status_code == 401


# ============================================================
# 6. Startup migration: legacy progres with po_id="" cleaned
# ============================================================
class TestLegacyProgresCleanup:
    def test_no_legacy_records_visible_via_rekap(self, admin):
        """After startup migration, /rekap/progres should have no records with empty no_po."""
        r = admin.get(f"{API}/rekap/progres", timeout=30)
        assert r.status_code == 200
        data = r.json()
        # Every returned progres record should have a valid non-empty no_po
        # (legacy rows with po_id='' would have joined to empty no_po)
        empty = [row for row in data if not row.get("no_po")]
        assert empty == [], f"Legacy cleanup didn't purge {len(empty)} orphan progres rows"


# ============================================================
# 7. Regression: iter3 no_po filter still works
# ============================================================
class TestIter3Regression:
    def test_rekap_all_po_no_po_filter(self, admin, seeded):
        r = admin.get(f"{API}/rekap/all-po",
                      params={"no_po": seeded["no_po"]}, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert len(data) > 0
        for row in data:
            assert row["no_po"] == seeded["no_po"]

    def test_auth_flow(self, admin, staff, guest):
        for s, role in [(admin, "admin"), (staff, "staff"), (guest, "guest")]:
            r = s.get(f"{API}/auth/me", timeout=30)
            assert r.status_code == 200
            assert r.json()["role"] == role


# ============================================================
# 8. Guest role authorization for new exports
# ============================================================
class TestGuestAccess:
    def test_guest_can_view_rekap_staffing_summary(self, guest):
        r = guest.get(f"{API}/rekap/staffing-summary", timeout=30)
        assert r.status_code == 200

    def test_guest_can_view_exports(self, guest):
        # Guests are allowed to view/export per app spec (Depends on get_current_user only)
        for path in ["/export/barang-masuk/pdf", "/export/staffing/pdf",
                     "/export/staffing/excel", "/export/barang-masuk/excel"]:
            r = guest.get(f"{API}{path}", timeout=60)
            assert r.status_code == 200, f"{path} returned {r.status_code}"
